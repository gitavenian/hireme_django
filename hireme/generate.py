import openpyxl
from random import randint, choice

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active

# Define column headers
headers = ["id", "emp_skill", "emp_exp", "required_skill", "required_exp", "convenient"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

# Job fields
job_fields = {
    "IT and Engineering": [
        "Front End Developer", "Back End Developer", "Full Stack Developer", "Flutter Developer", 
        "AI Development", "UX/UI Designer", "Graphic Designer", "IT Engineer", "Solutions Architect",
        "Game Developer", "Mechanical Engineer", "Electrical Engineer", "Civil Engineer",
        "Chemical Engineer", "Biomedical Engineer", "Environmental Engineer"
    ],
    "Healthcare": [
        "Medical Doctor", "Nurse", "Pharmacist", "Physiotherapist", "Veterinarian", "Dentist",
        "Lab Technician"
    ],
    "Arts and Entertainment": [
        "Jeweler", "Makeup Artist", "Hairdresser / Barber", "Artist"
    ],
    "Hospitality and Services": [
        "Chef", "Barista / Waitor", "Tour Guide", "Event Planner", "Hotel / Restaurant Manager"
    ],
    "Sports and Fitness": [
        "Fitness Coach", "Personal Trainer", "Yoga Instructor", "Sports Therapist"
    ],
    "Marketing and Management": [
        "Marketing Manager", "Manager", "Project Manager", "Business Analyst", "Human Resources Manager"
    ],
    "Law and Order": [
        "Lawyer", "Judge"
    ],
    "Academia and Education": [
        "Teacher", "Professor"
    ],
    "Miscellaneous": [
        "Cybersecurity Analyst", "Technical Writer", "Florist", "Machinist", "Electrician", "Pilot",
        "Translator"
    ]
}

# Flatten the skills list for random choices
all_skills = [skill for sublist in job_fields.values() for skill in sublist]

# Fill data
for i in range(2, 502):  # 500 rows
    # Fill ID
    ws.cell(row=i, column=1, value=i - 1)

    # Choose a random job field for emp_skill and required_skill
    emp_skill = choice(all_skills)
    required_skill = choice(all_skills)

    # Fill experience
    emp_exp = randint(0, 5)
    required_exp = randint(0, 5)

    # Determine convenient
    convenient = 1 if emp_skill == required_skill and emp_exp >= required_exp else 0

    # Ensure emp_skill and required_skill are not the same and convenient is 1 in some cases
    if convenient == 0:
        # Randomly decide to make convenient 1
        if randint(0, 1) == 1:
            emp_skill = required_skill
            emp_exp = max(emp_exp, required_exp)
            convenient = 1

    ws.cell(row=i, column=2, value=emp_skill)
    ws.cell(row=i, column=3, value=emp_exp)
    ws.cell(row=i, column=4, value=required_skill)
    ws.cell(row=i, column=5, value=required_exp)
    ws.cell(row=i, column=6, value=convenient)

# Save the workbook
wb.save("employee_data.xlsx")